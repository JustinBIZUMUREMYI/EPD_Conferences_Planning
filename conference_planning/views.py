
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect, get_object_or_404
from .forms import registerForm, LoginForm, PartnerForm, SponsorForm, SpeakerForm, PanalistForm, TestimonialForm, boothForm, VideoForm, PhotoForm, AgendaForm, EventForm,RegisterdayForm, SponsorshipForm,PDFFileForm,BookSponsorshipForm,PDFFileForm, BookAccessoryForm, BookBoothForm, InternsForm 
from django.contrib import messages
from .models import Attendees, Partner, Sponsor, Speaker, Event, Agenda, Panalist, booth, Testimonial, PreviousVideos, PreviousPhotos, Agenda, Event, Event_days, Sponsorships,Document,BookSponsorship, Countdown, Document, FloorPlan,accessory,Booth_space, BookBooth,BookAccessory,PreviousConferences, Interns, internship_document  
from django.urls import reverse
from django.views.generic import ListView, CreateView, UpdateView, TemplateView, DetailView
from django.views.generic.edit import FormView
from .forms import PartnerForm
from django.urls import reverse_lazy
from django.core.exceptions import ValidationError
import openpyxl
from django.http import HttpResponse
from datetime import datetime




def index(request):

    # fetching the attendees and their categories as well
    total_attendees = Attendees.objects.all()
    attendees = total_attendees.count()
    locals = Attendees.objects.filter(attendee_type='Local').count()
    internationals = Attendees.objects.filter(attendee_type='International').count()
    students = Attendees.objects.filter(attendee_type='Student').count()
    
    # fetching the statistics
    total_sponsors = Sponsor.objects.all().count()
    total_partners = Partner.objects.all().count()
    total_panalists = Panalist.objects.all().count()
    total_speakers = Speaker.objects.all().count()
    list_videos = PreviousVideos.objects.filter(title='Higlight')\
    

    # fetching testimonials
    testimonials = Testimonial.objects.all()

    # fetching sponsors    
    sponsors = Sponsor.objects.all()

    # counting down timer

    countdown = Countdown.objects.first()  # Assuming you only have one countdown


    context = {'number_attendees': attendees, 
             'locals_number': locals, 
             'internationals_number':internationals, 
             'students_number': students,
             'testimonials' : testimonials, 
             'sponsor_number': total_sponsors,
             'partner_number': total_partners,
             'panalist_number': total_panalists,
             'speaker_number': total_speakers,
             'end_time': countdown.end_time.isoformat(),
              'list_videos': list_videos,
            'sponsors': sponsors
             }
    return render(request, 'conference_planning/index.html', context)


def agenda(request):
    pdf = Document.objects.filter(title = 'agenda')
    context = {
        'documents': pdf
    }
    return render(request, 'conference_planning/schedule.html', context)


def speakers(request):
    list_speakers = Speaker.objects.all()
    list_panelists = Panalist.objects.all()

    context = {
        'speakers_list': list_speakers,
        'panelists_list': list_panelists
    }
    return render(request, 'conference_planning/speakers.html', context)

def speakers_details(request, speaker_id):
    speaker = get_object_or_404(Speaker, id=speaker_id)
    return render(request, 'conference_planning/speakers_details.html', {'speaker': speaker})

def panelist_details(request, panelist_id):
    panelist = get_object_or_404(Panalist, id=panelist_id)
    return render(request, 'conference_planning/panelist_details.html', {'panelist': panelist})

def sponsors(request):
    list_sponsors = Sponsor.objects.all()
    context = {
        'sponsors_list': list_sponsors
    }
    return render(request, 'conference_planning/sponsors.html', context)

def official_partners(request):
    list_officialpartners = Partner.objects.all()
    context = {
        'partners_list': list_officialpartners
    }
    return render(request, 'conference_planning/official_partners.html', context)

def sponsorship_packages(request):

    pdf = Document.objects.filter(title = 'sponsorships')
    sponsorships = Sponsorships.objects.all()
    if request.method == 'POST':
        form = BookSponsorshipForm(request.POST)
        if form.is_valid():
            form.save()
            success_message = "Thank you for registering, Kindly check your email for further information regarding your sponsorship. "
            messages.success(request, success_message)
            return redirect('sponsorships')
            
            
    else:
        form = BookSponsorshipForm()
    context = {
        'sponsorships': sponsorships,
        'documents': pdf,
        'form': form
        
    }

    return render(request, 'conference_planning/sponsorship_packages.html', context)




# def exhbition(request):
#     pdf = Document.objects.filter(title = 'exhibition')
#     floorplan = FloorPlan.objects.all()
#     exhibition = booth.objects.all()
#     accessories = accessory.objects.all()

#     if request.method == 'POST':
#         form_Booth = BookBoothForm(request.POST)
#         from_Accessory = BookAccessoryForm(request.POST)

        

#         if form_Booth.is_valid():
#             form_Booth.save()
#             success_message = "Thank you for booking, Kindly check your email for further information regarding exhibition. "
#             messages.success(request, success_message)
#             return redirect('exhbition')

#         elif from_Accessory.is_valid():
#               from_Accessory.save()
#               success_message = "Thank you for booking, Kindly check your email for further information regarding your exhibition. "
#               messages.success(request, success_message)
#               return redirect('exhbition')
            
#     else:
#         form_Booth = BookBoothForm()
#         form_Accessory = BookAccessoryForm()


#     context = {
#         'booths':  exhibition,
#          'documents': pdf,
#          'floorplan':floorplan,
#          'accessories':accessories,
#          'form_booth': form_Booth,
#          'form_accessory': form_Accessory

#     }
#     return render(request, 'conference_planning/exhbitions.html', context)


def exhbition(request):
    pdf = Document.objects.filter(title='exhibition')
    floorplan = FloorPlan.objects.all()
    exhibition = booth.objects.all()
    accessories = accessory.objects.all()

    if request.method == 'POST':
        booth_id = request.POST.get('booth_id')
        booth_space = request.POST.get('booth_space')

        accessory_id = request.POST.get('accessory_id')
        accessory_name = request.POST.get('accessory_name')

        form_Booth = BookBoothForm(request.POST)
        form_Accessory = BookAccessoryForm(request.POST)

        if form_Booth.is_valid():
            booking = form_Booth.save(commit=False)
            if booth_id:
                booking.booth = booth.objects.get(id=booth_id)
            if booth_space:
                booking.booth_space = booth_space
            booking.save()
            messages.success(request, "Thank you for booking a booth.")
            return redirect('exhbition')

        elif form_Accessory.is_valid():
            booking = form_Accessory.save(commit=False)
            if accessory_id:
                booking.accessory = accessory.objects.get(id=accessory_id)
            if accessory_name:
                booking.accessory_name = accessory_name  # if applicable
            booking.save()
            messages.success(request, "Thank you for booking an accessory.")
            return redirect('exhbition')

    else:
        form_Booth = BookBoothForm()
        form_Accessory = BookAccessoryForm()

    context = {
        'booths': exhibition,
        'documents': pdf,
        'floorplan': floorplan,
        'accessories': accessories,
        'form_booth': form_Booth,
        'form_accessory': form_Accessory,
    }
    return render(request, 'conference_planning/exhbitions.html', context)

    


def csr(request):
    pdf = Document.objects.filter(title = 'lighting')
    context = {
        'documents': pdf
    }
    return render(request, 'conference_planning/csr.html', context)

def First_edition(request):
    pdf = Document.objects.filter(title = 'first_edition')
    context = {
        'documents': pdf
    }
    return render(request, 'conference_planning/editions/firstedition.html', context)

def Second_edition(request):
    pdf = Document.objects.filter(title = 'second_edition')
    context = {
        'documents': pdf
    }
    return render(request, 'conference_planning/editions/secondedition.html', context)

def Third_edition(request):
    pdf = Document.objects.filter(title = 'third_edition')
    context = {
        'documents': pdf
    }
    return render(request, 'conference_planning/editions/thirdeditions.html', context)
    

def local_registration(request):

    # if request.method == 'POST':
    #         form = registerForm(request.POST)
    #         if form.is_valid():     
    #             if form.cleaned_data['country_code'] == '250':
    #                 attendee_type = 'Local'
                
    #             # if 'student_number' in form.cleaned_data and form.cleaned_data['student_number']:
    #             #     stud_number = form.cleaned_data['student_number']
    #             #     attendee_type = 'Student'
                    
    #             # if 'university' in form.cleaned_data and form.cleaned_data['university']:
    #             #     school = form.cleaned_data['university']
    #             #     attendee_type = 'Student'
    #             # validating phone number to accept only digits 
    #             # phone = form.cleaned_data['phone']
    #             # if not phone.isdigit():
    #             #     form.add_error('phone', 'Phone number must contain only digits.')
    #             #     return_url = form.cleaned_data['return_url']
    #             #     return render(request, return_url, {'form': form})
    #             # Check for existing attendee
    #             if Attendees.objects.filter(
    #                 identity=form.cleaned_data['identity']
    #             ).exists() or Attendees.objects.filter(
    #                 email=form.cleaned_data['email']
    #             ).exists():
    #                 error_message = "You are already registered, Please click back to home."
    #                 return_url = form.cleaned_data['return_url']
    #                 return render(request, return_url, {'error': error_message, 'form': form})
                
                
    #             attendee = Attendees(
    #                 names=form.cleaned_data['name'],
    #                 identity=form.cleaned_data['identity'],
    #                 email=form.cleaned_data['email'],
    #                 phone=form.cleaned_data['country_code'] + form.cleaned_data['phone'],
    #                 attendee_type=attendee_type,
    #                 # university=school,
    #                 # student_number=stud_number,
    #                 category=form.cleaned_data['category'],
    #                 country=form.cleaned_data['country'],
    #                 organization=form.cleaned_data['organization']
    #             )
    #             attendee.save()
    #             success_message = (
    #                 "Greetings from Energy Private Developers Association (EPD), "
    #                 "Thank you for registering to be part of Renewable Energy Week Conference Sept 2024. "
    #                 "We are pleased to inform you that you have completed the first step (1/2) of the registration process. "
    #                 "Within the next 24 hours, you will receive an email from us containing all the necessary details regarding the payment to complete your registration. "
    #                 "We appreciate your step and look forward to hosting you. Best regards, EPD"
    #             )
    #             redirect_url = request.META.get('HTTP_REFERER', '/')
    #             messages.success(request, success_message)
    #             return redirect(redirect_url)
    #         else:
    #             # return_url = form.cleaned_data['return_url']
    #             # return render(request, return_url, {'form': form})
    #             return render(request, 'conference_planning/registration/registration_local.html')
    # else:
    #     form = registerForm()
    #     return render(request, 'conference_planning/registration/registration_local.html', {'form':form})



    return render(request, 'conference_planning/registration/registration_local.html')


def international_registration(request):
    message = request.GET.get('message', None)
    context = {}
    if message:
        context['message'] = message
    return render(request, 'conference_planning/registration/registration_international.html', context)


def students_registration(request):
    return render(request, 'conference_planning/registration/registration_students.html')




# def register(request):
#     if request.method == 'POST':
#         form = registerForm(request.POST)
#         if form.is_valid():
#             attendee_type = 'Local'
#             stud_number = None
#             school = None
            
#             if form.cleaned_data['country_code'] != '250':
#                 attendee_type = 'International'
            
#             if 'student_number' in form.cleaned_data and form.cleaned_data['student_number']:
#                 stud_number = form.cleaned_data['student_number']
#                 attendee_type = 'Student'
                
#             if 'university' in form.cleaned_data and form.cleaned_data['university']:
#                 school = form.cleaned_data['university']
#                 attendee_type = 'Student'
#             # validating phone number to accept only digits 
#             # phone = form.cleaned_data['phone']
#             # if not phone.isdigit():
#             #     form.add_error('phone', 'Phone number must contain only digits.')
#             #     return_url = form.cleaned_data['return_url']
#             #     return render(request, return_url, {'form': form})
#             # Check for existing attendee
#             if Attendees.objects.filter(
#                 identity=form.cleaned_data['identity']
#             ).exists() or Attendees.objects.filter(
#                 email=form.cleaned_data['email']
#             ).exists():
#                 error_message = "You are already registered, Please click back to home."
#                 return_url = form.cleaned_data['return_url']
#                 return render(request, return_url, {'error': error_message, 'form': form})
               
            
#             attendee = Attendees(
#                 names=form.cleaned_data['name'],
#                 identity=form.cleaned_data['identity'],
#                 email=form.cleaned_data['email'],
#                 phone=form.cleaned_data['country_code'] + form.cleaned_data['phone'],
#                 attendee_type=attendee_type,
#                 university=school,
#                 student_number=stud_number,
#                 category=form.cleaned_data['category'],
#                 country=form.cleaned_data['country'],
#                 organization=form.cleaned_data['organization']
#             )
#             attendee.save()
#             success_message = (
#                 "Greetings from Energy Private Developers Association (EPD), "
#                 "Thank you for registering to be part of Renewable Energy Week Conference Sept 2024. "
#                 "We are pleased to inform you that you have completed the first step (1/2) of the registration process. "
#                 "Within the next 24 hours, you will receive an email from us containing all the necessary details regarding the payment to complete your registration. "
#                 "We appreciate your step and look forward to hosting you. Best regards, EPD"
#             )
#             redirect_url = request.META.get('HTTP_REFERER', '/')
#             messages.success(request, success_message)
#             return redirect(redirect_url)
#         else:
#             form = registerForm()
#             return render(request, return_url, {'form': form})
#     else:
#         form = registerForm()
#         return render(request, return_url, {'form': form})



# the current register codes.


# def register(request):
#     if request.method == 'POST':
#         form = registerForm(request.POST)
#         if form.is_valid():
#             attendee_type = 'Local'
#             stud_number = None
#             school = None
            
#             if form.cleaned_data['country_code'] != '250':
#                 attendee_type = 'International'
            
#             if 'student_number' in form.cleaned_data and form.cleaned_data['student_number']:
#                 stud_number = form.cleaned_data['student_number']
#                 attendee_type = 'Student'
                
#             if 'university' in form.cleaned_data and form.cleaned_data['university']:
#                 school = form.cleaned_data['university']
#                 attendee_type = 'Student'
            
#             attendee = Attendees(
#                 names=form.cleaned_data['name'],
#                 identity=form.cleaned_data['identity'],
#                 email=form.cleaned_data['email'],
#                 phone=form.cleaned_data['country_code'] + form.cleaned_data['phone'],
#                 attendee_type=attendee_type,
#                 university=school,
#                 student_number=stud_number,
#                 category = form.cleaned_data['category'],
#                 country=form.cleaned_data['country'],
#                 organization=form.cleaned_data['organization']
#             )
#             attendee.save()
#             success_message = """
#             Greetings from Energy Private Developers Association (EPD),<br><br>

#             Thank you for registering to attend the Renewable Energy Week Conference in September 2025. You have successfully completed the first step (1/2) of the registration process.<br><br>

#             <strong>To complete your registration, please proceed with the payment using one of the options below:</strong><br><br>

#             <strong>💳 ECOBANK RWANDA</strong><br>
#             Account Name: Energy Private Developers<br>
#             USD Account: 6775009645<br>
#             RWF Account: 6775008215<br>
#             SWIFT Code: ECOCRWRWXX<br>
#             Branch: Head Office, Avenue de la Paix, PO Box 3268, Kigali-Rwanda<br><br>

#             <strong>🏦 Bank of Kigali (USD)</strong><br>
#             Account Name: Energy Private Developers Association<br>
#             Account Number: 100188383844<br>
#             IBAN: RW43040100188383844840<br>
#             SWIFT Code: BKIGRWRWXXX<br>
#             Branch: Head Office<br><br>

#             <strong>🏦 Bank of Kigali (RWF)</strong><br>
#             Account Name: Energy Private Developers Association<br>
#             Account Number: 100188383097<br>
#             IBAN: RW27040100188383097646<br>
#             SWIFT Code: BKIGRWRWXXX<br>
#             Branch: Head Office<br><br>

#             <strong>📱 MoMo Pay</strong><br>
#             Code: *182*8*1*077863#<br>
#             Name: Energy Private Developers Association<br><br>

#             Once the payment is made, please keep your proof of payment. You will receive a confirmation email from us within the next 24 hours.<br><br>

#             We appreciate your participation and look forward to hosting you!<br><br>
#             Best regards,<br>
#             EPD Team
#             """
#             redirect_url = request.META.get('HTTP_REFERER', '/')
#             messages.success(request, success_message)
#             return redirect(redirect_url)
#         else:
#             return_url = form.cleaned_data['return_url']
#             return render(request, return_url, {'form': form})
#     else:
#         form = registerForm()
    
#     return render(request, return_url, {'form': form})


# Trying to send the authomatic email
from django.core.mail import EmailMessage
from django.utils.html import strip_tags

def register(request):
    if request.method == 'POST':
        form = registerForm(request.POST)
        if form.is_valid():
            attendee_type = 'Local'
            stud_number = None
            school = None
            
            if form.cleaned_data['country_code'] != '250':
                attendee_type = 'International'
            
            if 'student_number' in form.cleaned_data and form.cleaned_data['student_number']:
                stud_number = form.cleaned_data['student_number']
                attendee_type = 'Student'
                
            if 'university' in form.cleaned_data and form.cleaned_data['university']:
                school = form.cleaned_data['university']
                attendee_type = 'Student'
            
            attendee = Attendees(
                names=form.cleaned_data['name'],
                identity=form.cleaned_data['identity'],
                email=form.cleaned_data['email'],
                phone=form.cleaned_data['country_code'] + form.cleaned_data['phone'],
                attendee_type=attendee_type,
                university=school,
                student_number=stud_number,
                category=form.cleaned_data['category'],
                country=form.cleaned_data['country'],
                organization=form.cleaned_data['organization']
            )
            attendee.save()

            # Compose email message
            html_message = """
            Greetings from Energy Private Developers Association (EPD),<br><br>

            Thank you for registering to attend the Renewable Energy Week Conference in September 2025. You have successfully completed the first step (1/2) of the registration process.<br><br>

            <strong>To complete your registration, please proceed with the payment using one of the options below:</strong><br><br>

            <strong>💳 ECOBANK RWANDA</strong><br>
            Account Name: Energy Private Developers<br>
            USD Account: 6775009645<br>
            RWF Account: 6775008215<br>
            SWIFT Code: ECOCRWRWXX<br>
            Branch: Head Office, Avenue de la Paix, PO Box 3268, Kigali-Rwanda<br><br>

            <strong>🏦 Bank of Kigali (USD)</strong><br>
            Account Name: Energy Private Developers Association<br>
            Account Number: 100188383844<br>
            IBAN: RW43040100188383844840<br>
            SWIFT Code: BKIGRWRWXXX<br>
            Branch: Head Office<br><br>

            <strong>🏦 Bank of Kigali (RWF)</strong><br>
            Account Name: Energy Private Developers Association<br>
            Account Number: 100188383097<br>
            IBAN: RW27040100188383097646<br>
            SWIFT Code: BKIGRWRWXXX<br>
            Branch: Head Office<br><br>

            <strong>📱 MoMo Pay</strong><br>
            Code: *182*8*1*077863#<br>
            Name: Energy Private Developers Association<br><br>

            Once the payment is made, please keep your proof of payment. You will receive a confirmation email from us within the next 24 hours.<br><br>

            We appreciate your participation and look forward to hosting you!<br><br>
            Best regards,<br>
            EPD Team
            """

            plain_message = strip_tags(html_message)  # fallback for non-HTML email readers
            recipient = form.cleaned_data['email']

            email = EmailMessage(
                subject='EPD Conference 2025 Registration Confirmation',
                body=plain_message,
                from_email=None,  # Uses DEFAULT_FROM_EMAIL
                to=[recipient],
            )
            email.content_subtype = 'html'
            email.body = html_message
            email.send()

            # Show success message on website
            messages.success(request, "Thank you for registering! A confirmation email has been sent to your inbox.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            return_url = form.cleaned_data['return_url']
            return render(request, return_url, {'form': form})
    else:
        form = registerForm()

    return render(request, return_url, {'form': form})






def login_view(request):
    return render(request, 'conference_planning/administration/login.html')


def auth(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('administration:dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    return render(request, 'conference_planning/administration/login.html', {'form': form})


@login_required
def dashboard(request):
    attendees_list = Attendees.objects.all()
    partner_total = Partner.objects.all().count()
    sponsor_total = Sponsor.objects.all().count()
    speaker_total = Speaker.objects.all().count()
    panalist_total = Panalist.objects.all().count()
    interns_total = Interns.objects.all().count()
    locals = Attendees.objects.filter(attendee_type='Local').count()
    internationals = Attendees.objects.filter(attendee_type='International').count()
    students = Attendees.objects.filter(attendee_type='Student').count()
    counting_attendees = attendees_list.count()
    context = {'attendees_number':counting_attendees, 
               'locals_number': locals, 
               'internationals_number':internationals, 
               'students_number': students,
               'total_partner': partner_total,
               'total_sponsor':sponsor_total,
               'total_speaker':speaker_total,
               'total_panelist':panalist_total,
               'interns_total':interns_total,
               }
    return render(request, 'conference_planning/administration/index.html', context)


@login_required
def attendees(request):
    attendees_list = Attendees.objects.all().order_by('id')
    context = {'attendees': attendees_list}
    
    return render(request, 'conference_planning/administration/attends.html', context)


# the conference (previous videos and photos)
def previous_videos(request):
    list_videos = PreviousVideos.objects.all()
    context = {
        'videos_list': list_videos
    }
    return render(request, 'conference_planning/videos.html', context)
    
def previous_photos(request):
    list_photos = PreviousPhotos.objects.all()
    context = {
        'photos_list': list_photos
    }
    return render(request, 'conference_planning/photos.html', context)



# regustration forms
class registration_froms(TemplateView):
    template_name = 'conference_planning/registration/registration.html'

# Sponsors
class Register_sponsor(CreateView):
    form_class = SponsorForm
    success_url = reverse_lazy('administration:sponsor')
    template_name = 'conference_planning/administration/sponsor_registration.html'
    model = Sponsor
class sponsors_list(ListView):
    template_name = 'conference_planning/administration/sponsors_list.html'
    model = Sponsor
    context_object_name = 'sponsors'

    

# Partners
class Register_parner(CreateView):
    form_class = PartnerForm
    success_url = reverse_lazy('administration:partner')
    template_name = 'conference_planning/administration/partner_registration.html'
    model = Partner
class partners_list(ListView):
    template_name = 'conference_planning/administration/partners_list.html'
    model = Partner
    context_object_name = 'partners'


# speakers
class Register_speaker(CreateView):
    form_class = SpeakerForm
    success_url = reverse_lazy('administration:speaker')
    template_name = 'conference_planning/administration/speaker_registration.html'
    model = Speaker
class speakers_list(ListView):
    template_name = 'conference_planning/administration/speakers_list.html'
    model = Speaker
    context_object_name = 'speakers'

# Panalists
class Register_panalist(CreateView):
    form_class = PanalistForm
    success_url = reverse_lazy('administration:panalist')
    template_name = 'conference_planning/administration/panalist_registration.html'
    model = Panalist
class panalists_list(ListView):
    template_name = 'conference_planning/administration/panalists_list.html'
    model = Panalist
    context_object_name = 'panalists'
    
# Testimonials
class Register_testimonial(CreateView):
    form_class = TestimonialForm
    success_url = reverse_lazy('administration:testimonial')
    template_name = 'conference_planning/administration/testimonial_registration.html'
    model = Testimonial
class testimonials_list(ListView):
    template_name = 'conference_planning/administration/testimonials_list.html'
    model = Testimonial
    context_object_name = 'testimonials'

# booths
class Register_booth(CreateView):
    form_class = boothForm
    success_url = reverse_lazy('administration:booth')
    template_name = 'conference_planning/administration/booth_registration.html'
    model = booth
class booths_list(ListView):
    template_name = 'conference_planning/administration/booths_list.html'
    model = booth
    context_object_name = 'booths'


# agenda
class Register_agenda(CreateView):
    form_class = AgendaForm
    success_url = reverse_lazy('administration:agendas')
    template_name = 'conference_planning/administration/register_agenda.html'
    model = Agenda
class Display_agenda(ListView):
    template_name = 'conference_planning/administration/display_agenda.html'
    model = Agenda
    context_object_name = 'agendas'   
    
# Event
class Register_event(CreateView):
    form_class = EventForm
    success_url = reverse_lazy('administration:event')
    template_name = 'conference_planning/administration/register_event.html'
    model = Event
class Display_event(ListView):
    template_name = 'conference_planning/administration/display_event.html'
    model = Event
    context_object_name = 'events'  

class Register_Eventday(CreateView):
    form_class = RegisterdayForm 
    success_url = reverse_lazy('administration:dayofevent')
    template_name = 'conference_planning/administration/register_dayevent.html'
    model = Event_days


# previous conferences videos and photos
class videos(CreateView):
    form_class = VideoForm
    success_url = reverse_lazy('administration:video')
    template_name = 'conference_planning/administration/previous_videos.html'
    model =  PreviousVideos
class photos(CreateView):
    form_class = PhotoForm
    success_url = reverse_lazy('administration:photo')
    template_name = 'conference_planning/administration/previous_photos.html'
    model =  PreviousPhotos

# sponsorship packages  
class sponsorship(CreateView):
    form_class = SponsorshipForm
    success_url = reverse_lazy('administration:sponsorships')
    template_name = 'conference_planning/administration/sponsorships.html'
    model = Sponsorships

# uploading a document
class uploadPDF(CreateView):
    form_class = PDFFileForm
    success_url = reverse_lazy('administration:sponsorships')
    template_name =  'conference_planning/administration/sponsorships.html'
    model = Document

class PDFFile(ListView):
    model = Document
    template_name = 'conference_planning/sponsorship_packages.html'
    context_object_name = 'pdf_document'

# Locals list
@login_required
def locals(request):
    local_list = Attendees.objects.filter(attendee_type='Local').order_by('id')
    locals_attends = Attendees.objects.filter(attendee_type='Local').count()
    context = {
        'locals': local_list,
        'total_locals': locals_attends
    
    }
    
    return render(request, 'conference_planning/administration/locals_list.html', context)


# internations list
@login_required
def internationals(request):
    international_list = Attendees.objects.filter(attendee_type='International').order_by('id')
    total_international = Attendees.objects.filter(attendee_type='International').count()
    context = {'internationals': international_list, 'total_internationls': total_international}
    
    return render(request, 'conference_planning/administration/internationals_list.html', context)


# students list
@login_required
def students(request):
    student_list = Attendees.objects.filter(attendee_type='Student').order_by('id')
    context = {'students': student_list}
    
    return render(request, 'conference_planning/administration/students_list.html', context)




# reports

# exporting all the registered attends 
def export_attendees_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=attendees.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendees"

    # Add column headers
    columns = ['Name', 'Identity', 'Email', 'Category', 'Country', 'Organization']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    attendees = Attendees.objects.all()
    for attendee_num, attendee in enumerate(attendees, 2):
        ws.cell(row=attendee_num, column=1, value=attendee.names)
        ws.cell(row=attendee_num, column=2, value=attendee.identity)
        ws.cell(row=attendee_num, column=3, value=attendee.email)
        ws.cell(row=attendee_num, column=4, value=attendee.category)
        ws.cell(row=attendee_num, column=5, value=attendee.country)
        ws.cell(row=attendee_num, column=6, value=attendee.organization)

    wb.save(response)
    return response


# exporting locals  

def export_locals_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Locals.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Local Delegates"

    # Add column headers
    columns = ['Name', 'Identity', 'Phone','Email', 'Category', 'Country', 'Organization']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    attendees = Attendees.objects.filter(attendee_type='Local')
    for attendee_num, attendee in enumerate(attendees, 2):
        ws.cell(row=attendee_num, column=1, value=attendee.names)
        ws.cell(row=attendee_num, column=2, value=attendee.identity)
        ws.cell(row=attendee_num, column=3, value=attendee.phone)
        ws.cell(row=attendee_num, column=4, value=attendee.email)
        ws.cell(row=attendee_num, column=5, value=attendee.category)
        ws.cell(row=attendee_num, column=6, value=attendee.country)
        ws.cell(row=attendee_num, column=7, value=attendee.organization)

    wb.save(response)
    return response



# exporting Internationls 
def export_internationals_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Internationals.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "International Delegates"

    # Add column headers
    columns = ['Name', 'Identity', 'Phone','Email', 'Category', 'Country', 'Organization']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    attendees = Attendees.objects.filter(attendee_type='International')
    for attendee_num, attendee in enumerate(attendees, 2):
        ws.cell(row=attendee_num, column=1, value=attendee.names)
        ws.cell(row=attendee_num, column=2, value=attendee.identity)
        ws.cell(row=attendee_num, column=3, value=attendee.phone)
        ws.cell(row=attendee_num, column=4, value=attendee.email)
        ws.cell(row=attendee_num, column=5, value=attendee.category)
        ws.cell(row=attendee_num, column=6, value=attendee.country)
        ws.cell(row=attendee_num, column=7, value=attendee.organization)

    wb.save(response)
    return response


# exporting the list of the speakers 
def export_speakers_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Current_speakers_list.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "the Speakers of the conference"

    # Add column headers
    columns = ['Name', 'Title','Email']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    attendees = Speaker.objects.all()
    for attendee_num, attendee in enumerate(attendees, 2):
        ws.cell(row=attendee_num, column=1, value=attendee.names)
        ws.cell(row=attendee_num, column=2, value=attendee.title)
        ws.cell(row=attendee_num, column=3, value=attendee.email)


    wb.save(response)
    return response


# exporting the list of the Panelist 
def export_panelists_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Current_Panelists_list.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "the Speakers of the conference"

    # Add column headers
    columns = ['Name', 'Title','Email']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    attendees = Panalist.objects.all()
    for attendee_num, attendee in enumerate(attendees, 2):
        ws.cell(row=attendee_num, column=1, value=attendee.names)
        ws.cell(row=attendee_num, column=2, value=attendee.title)
        ws.cell(row=attendee_num, column=3, value=attendee.email)


    wb.save(response)
    return response



 
class Register_Interns(CreateView):
    form_class = InternsForm
    success_url = reverse_lazy('administration:internship_Application')
    template_name = 'conference_planning/register_interns.html'
    model = Interns

# class interns_list(TemplateView):
#    template_name = 'conference_planning/register_interns.html'

#    def get_context_data(self, **kwargs):


#     context = super().get_context_data(**kwargs)

#         # Adding two different contexts
#     context['members_list'] = Testimonial.objects.all()  # First context
#     context['documents'] = Document.objects.filter(title='AESG')  # Second context
        
#     return context

class interns_list(ListView):
    model = internship_document  # Define the model that should be listed
    template_name = 'conference_planning/register_interns.html'
    context_object_name = 'interns'  
    ordering = ['-id']
    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     # Adding two additional contexts
    #     context['members_list'] = Testimonial.objects.all()  # First context
    #     context['documents'] = Document.objects.filter(title='AESG')  # Second context
    #     return context


class submit_application(CreateView):
    form_class = InternsForm
    success_url = reverse_lazy('internship_Application')
    template_name = 'conference_planning/regitration_form.html'
    model = Interns

    def form_valid(self, form):
      
        response = super().form_valid(form)
        
        # Add a success message to be displayed after form submission
        messages.success(self.request, "Thank you for your application! We will be in touch with you shortly.")
        
        return response
        

def interns(request):
    interns_list = Interns.objects.all().order_by('id')
    context = {'interns': interns_list}
    
    return render(request, 'conference_planning/administration/internship.html', context)

  
# exporting the list of the interns
def export_applicants_to_excel(request):
    # Create an in-memory output file for the new workbook.
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Applicants.xlsx'

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "The list of Applicants"

    # Add column headers
    columns = ['Name', 'Phone Number','Email', 'University', 'Education Level', 'Qualification','Graduation Date','Campany','Degree','Resume','Additional Documents', 'Application Date']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    # Fetch data from the database
    interns = Interns.objects.all()
    for intern_num, intern in enumerate(interns, 2):
        ws.cell(row=intern_num, column=1, value=intern.Full_Name)
        ws.cell(row=intern_num, column=2, value=intern.Phone)
        ws.cell(row=intern_num, column=3, value=intern.Email)
        ws.cell(row=intern_num, column=4, value=intern.University)
        ws.cell(row=intern_num, column=5, value=intern.Education_level)
        ws.cell(row=intern_num, column=6, value=intern.Qualification)
        ws.cell(row=intern_num, column=7, value=intern.Graduation_date)

        ws.cell(row=intern_num, column=8, value=str(intern.Host_Company))
        ws.cell(row=intern_num, column=9, value=intern.Degree.name if intern.Degree else 'No file uploaded')
        
        # Get Resume and Additional Documents Names
        ws.cell(row=intern_num, column=10, value=intern.Resume.name if intern.Resume else 'No file uploaded')
        ws.cell(row=intern_num, column=11, value=intern.Other_documents.name if intern.Other_documents else 'No file uploaded')

        # Handle registered_on (checking if it's timezone-aware)
        if isinstance(intern.registered_on, datetime):
            if intern.registered_on.tzinfo is not None:  # If timezone-aware
                registered_on_naive = intern.registered_on.astimezone().replace(tzinfo=None)
            else:
                registered_on_naive = intern.registered_on
            ws.cell(row=intern_num, column=12, value=registered_on_naive.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            ws.cell(row=intern_num, column=12, value='')

    wb.save(response)
    return response











# # exporting the list of speakers
# def export_speakers_to_excel(request):
#     # Create an in-memory output file for the new workbook.
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Speakers.xlsx'

#     # Create a workbook and add a worksheet.
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "International Delegates"

#     # Add column headers
#     columns = ['Name', 'Phone','Email', 'Title']
#     for col_num, column_title in enumerate(columns, 1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.value = column_title

#     # Fetch data from the database
#     attendees = Attendees.objects.filter(attendee_type='International')
#     for attendee_num, attendee in enumerate(attendees, 2):
#         ws.cell(row=attendee_num, column=1, value=attendee.names)
#         ws.cell(row=attendee_num, column=2, value=attendee.identity)
#         ws.cell(row=attendee_num, column=3, value=attendee.phone)
#         ws.cell(row=attendee_num, column=4, value=attendee.email)
#         ws.cell(row=attendee_num, column=5, value=attendee.category)
#         ws.cell(row=attendee_num, column=6, value=attendee.country)
#         ws.cell(row=attendee_num, column=7, value=attendee.organization)

#     wb.save(response)
#     return response



# # exporting the list panelist
# def export_panelists_to_excel(request):
#     # Create an in-memory output file for the new workbook.
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Panelists.xlsx'

#     # Create a workbook and add a worksheet.
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "International Delegates"

#     # Add column headers
#     columns = ['Name', 'Phone','Email', 'Title']
#     for col_num, column_title in enumerate(columns, 1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.value = column_title

#     # Fetch data from the database
#     attendees = Attendees.objects.filter(attendee_type='International')
#     for attendee_num, attendee in enumerate(attendees, 2):
#         ws.cell(row=attendee_num, column=1, value=attendee.names)
#         ws.cell(row=attendee_num, column=2, value=attendee.identity)
#         ws.cell(row=attendee_num, column=3, value=attendee.phone)
#         ws.cell(row=attendee_num, column=4, value=attendee.email)
#         ws.cell(row=attendee_num, column=5, value=attendee.category)
#         ws.cell(row=attendee_num, column=6, value=attendee.country)
#         ws.cell(row=attendee_num, column=7, value=attendee.organization)

#     wb.save(response)
#     return response


# # exporting the list of partners
# def export_Partners_to_excel(request):
#     # Create an in-memory output file for the new workbook.
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Partners.xlsx'

#     # Create a workbook and add a worksheet.
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Partners"

#     # Add column headers
#     columns = ['Name', 'Phone','Email', 'Title']
#     for col_num, column_title in enumerate(columns, 1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.value = column_title

#     # Fetch data from the database
#     attendees = Attendees.objects.filter(attendee_type='International')
#     for attendee_num, attendee in enumerate(attendees, 2):
#         ws.cell(row=attendee_num, column=1, value=attendee.names)
#         ws.cell(row=attendee_num, column=2, value=attendee.identity)
#         ws.cell(row=attendee_num, column=3, value=attendee.phone)
#         ws.cell(row=attendee_num, column=4, value=attendee.email)
#         ws.cell(row=attendee_num, column=5, value=attendee.category)
#         ws.cell(row=attendee_num, column=6, value=attendee.country)
#         ws.cell(row=attendee_num, column=7, value=attendee.organization)

#     wb.save(response)
#     return response